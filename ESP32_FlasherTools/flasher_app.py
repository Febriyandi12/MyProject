# VERSI 1.0 - Log to Excel, Friendly Errors, Bug Fixes
import customtkinter as ctk
import serial
import serial.tools.list_ports
import subprocess
import threading
import sys
import os
from datetime import datetime
from tkinter import filedialog
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("="*50)
    print("PERINGATAN: Library 'openpyxl' tidak ditemukan.")
    print("Fitur 'Log to Excel' tidak akan berfungsi.")
    print("Silakan install dengan: pip install openpyxl")
    print("="*50)
    openpyxl = None # Set ke None agar program tetap jalan

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("ESP32 Flasher & Monitor (v4.1)") # Versi baru
        self.geometry("600x800") 
        
        self.grid_columnconfigure(0, weight=1)
        
        # --- Variabel State ---
        self.ser = None
        self.serial_running = False
        self.is_logging = False
        self.is_logging_paused = False
        self.log_data_buffer = [] 
        
        # --- Frame Konfigurasi ---
        config_frame = ctk.CTkFrame(self)
        config_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        config_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(config_frame, text="COM Port:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.com_port_menu = ctk.CTkOptionMenu(config_frame, values=["-"], command=self.on_port_select)
        self.com_port_menu.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        ctk.CTkLabel(config_frame, text="Baud Rate:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.baud_rate_menu = ctk.CTkOptionMenu(config_frame, values=["115200", "230400", "460800", "921600"])
        self.baud_rate_menu.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        self.baud_rate_menu.set("115200") 

        self.refresh_ports_button = ctk.CTkButton(config_frame, text="Refresh", width=80, command=self.refresh_ports)
        self.refresh_ports_button.grid(row=0, column=2, padx=10, pady=5)
        
        self.info_button = ctk.CTkButton(config_frame, text="Info", width=60, command=self.show_info_popup)
        self.info_button.grid(row=0, column=3, padx=10, pady=5) 

        # --- Frame Upload ---
        upload_frame = ctk.CTkFrame(self)
        upload_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        upload_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(upload_frame, text="Firmware:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.firmware_path = ctk.CTkEntry(upload_frame, placeholder_text="Pilih file APLIKASI (.bin)")
        self.firmware_path.grid(row=0, column=1, padx=(0,10), pady=5, sticky="ew")
        
        self.browse_button = ctk.CTkButton(upload_frame, text="...", width=40, command=self.browse_file)
        self.browse_button.grid(row=0, column=2, padx=(0,10), pady=5)

        self.flash_button = ctk.CTkButton(upload_frame, text="UPLOAD FIRMWARE", command=self.start_flash_thread)
        self.flash_button.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

        # --- Frame Erase ---
        erase_frame = ctk.CTkFrame(self)
        erase_frame.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        self.erase_button = ctk.CTkButton(erase_frame, text="ERASE FLASH (HAPUS TOTAL)", fg_color="red", hover_color="darkred", command=self.start_erase_thread)
        self.erase_button.pack(fill="x", padx=10, pady=10)

        self.progress_bar = ctk.CTkProgressBar(erase_frame, orientation="horizontal", indeterminate_speed=1)

        # --- Frame Serial Monitor ---
        monitor_frame = ctk.CTkFrame(self)
        monitor_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        monitor_frame.grid_columnconfigure(0, weight=1)
        monitor_frame.grid_rowconfigure(2, weight=1) # <-- [EDIT] Diubah ke row 2

        # --- [LABEL BARU 1] ---
        ctk.CTkLabel(monitor_frame, text="Serial Monitor & Logger (Output dari ESP32)", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(5,0), sticky="w")

        # --- Frame Kontrol Serial & Log ---
        monitor_controls = ctk.CTkFrame(monitor_frame, fg_color="transparent")
        monitor_controls.grid(row=1, column=0, columnspan=2, sticky="ew") # <-- [EDIT] Diubah ke row 1
        
        self.serial_connect_button = ctk.CTkButton(monitor_controls, text="Connect Serial", command=self.toggle_serial_connection, width=120)
        self.serial_connect_button.pack(side="left", padx=5, pady=5)
        
        self.serial_clear_button = ctk.CTkButton(monitor_controls, text="Clear", command=self.clear_serial_output, width=60)
        self.serial_clear_button.pack(side="left", padx=5, pady=5)

        self.log_start_button = ctk.CTkButton(monitor_controls, text="Start Log", command=self.toggle_log_start, fg_color="green", hover_color="darkgreen", width=90)
        self.log_start_button.pack(side="left", padx=5, pady=5)
        
        self.log_pause_button = ctk.CTkButton(monitor_controls, text="Pause Log", command=self.toggle_log_pause, state="disabled", width=90)
        self.log_pause_button.pack(side="left", padx=5, pady=5)

        self.log_stop_button = ctk.CTkButton(monitor_controls, text="Stop & Save", command=self.stop_and_save_log, state="disabled", fg_color="red", hover_color="darkred", width=100)
        self.log_stop_button.pack(side="left", padx=5, pady=5)
        
        # --- Sisa Serial Monitor ---
        self.serial_output = ctk.CTkTextbox(monitor_frame, state="disabled")
        self.serial_output.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew") # <-- [EDIT] Diubah ke row 2

        self.serial_input = ctk.CTkEntry(monitor_frame, placeholder_text="Kirim data...")
        self.serial_input.grid(row=3, column=0, padx=10, pady=(0,10), sticky="ew") # <-- [EDIT] Diubah ke row 3
        self.serial_input.bind("<Return>", self.send_serial_data) 

        self.serial_send_button = ctk.CTkButton(monitor_frame, text="Send", width=60, command=self.send_serial_data)
        self.serial_send_button.grid(row=3, column=1, padx=(0,10), pady=(0,10)) # <-- [EDIT] Diubah ke row 3

        # --- Frame Log Flasher ---
        # --- [EDIT] Dibuat frame baru untuk log bawah ---
        log_frame = ctk.CTkFrame(self)
        log_frame.grid(row=4, column=0, padx=10, pady=10, sticky="nsew")
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(1, weight=0) # <-- Biarkan ukurannya tetap

        # --- [LABEL BARU 2] ---
        ctk.CTkLabel(log_frame, text="Process Log (Output dari Software Flasher)", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, padx=10, pady=(5,0), sticky="w")
        
        self.log_textbox = ctk.CTkTextbox(log_frame, height=150, state="disabled") # <-- [EDIT] Dipindah ke log_frame
        self.log_textbox.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        
        # --- Label Kredit ---
        self.footer_label = ctk.CTkLabel(self, text="Developed by FEBRI R&D", font=ctk.CTkFont(size=10, slant="italic"), text_color="gray")
        self.footer_label.grid(row=5, column=0, padx=10, pady=(0,10), sticky="e") # <-- [EDIT] Diubah ke row 5
        
        # --- Konfigurasi row-weight ---
        self.grid_rowconfigure(3, weight=1) # Serial Monitor (bisa membesar)
        self.grid_rowconfigure(4, weight=0) # Process Log (ukuran tetap)
        
        self.refresh_ports()

    # --- [FUNGSI BARU] Parser Error ---
    def parse_esptool_error(self, error_log):
        if "Could not open port" in error_log or "Access is denied" in error_log:
            return "PORT SIBUK ATAU TIDAK ADA.\n\nPastikan port tidak dipakai aplikasi lain (misal: Serial Monitor) dan kabel terpasang dengan benar."
        if "Failed to connect to ESP32" in error_log:
            return "GAGAL TERHUBUNG KE ESP32.\n\nCoba tahan tombol 'BOOT' di board ESP32 saat proses upload dimulai."
        if "FileNotFoundError" in error_log:
             return f"FILE TIDAK DITEMUKAN.\n\n{error_log}"
        
        # Error umum
        return f"Upload Gagal. Cek log di bawah untuk detail teknis."

    # --- [FUNGSI BARU] Popup Notifikasi ---
    def show_popup(self, title, message):
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("450x200")
        popup.transient(self)
        popup.grab_set()
        popup.after(100, popup.lift)

        label = ctk.CTkLabel(popup, text=message, wraplength=410, font=ctk.CTkFont(size=14))
        label.pack(pady=20, padx=20, expand=True, fill="both")
        
        ok_button = ctk.CTkButton(popup, text="OK", command=popup.destroy, width=100)
        ok_button.pack(pady=10)

    # --- [FUNGSI BARU] Info Popup ---
    def show_info_popup(self):
        info_win = ctk.CTkToplevel(self)
        info_win.title("Cara Penggunaan")
        info_win.geometry("550x550") # Perbesar
        info_win.transient(self)
        info_win.grab_set()
        info_win.after(100, info_win.lift)

        textbox = ctk.CTkTextbox(info_win, wrap="word", font=ctk.CTkFont(size=13))
        textbox.pack(expand=True, fill="both", padx=10, pady=10)
        
        help_text = """
--- Tata Cara Penggunaan ---

1.  Pastikan ESP32 terhubung ke PC.
2.  Pilih 'COM Port' yang benar. Klik 'Refresh' jika tidak muncul.
3.  Baud Rate:
    -   Untuk 'UPLOAD' atau 'ERASE', software otomatis pakai baudrate tertinggi.
    -   Untuk 'Serial Monitor', sesuaikan dengan baud rate program Anda (misal: 115200).

--- Upload Firmware (Aplikasi) ---
1.  Klik '...' untuk memilih file APLIKASI (.bin) Anda.
2.  Klik 'UPLOAD FIRMWARE'.
3.  Software ini akan otomatis meng-upload:
    -   bootloader.bin (bawaan software)
    -   partitions.bin (bawaan software)
    -   File aplikasi .bin (pilihan Anda)
4.  PENTING: Jika Serial Monitor terhubung, software akan otomatis 'Disconnect' sebelum memulai upload.

--- Erase Flash ---
Tombol ini akan menghapus TOTAL data di ESP32.

--- Serial Monitor & Logger ---
1.  Pilih Baud Rate yang sesuai (misal: 115200).
2.  Klik 'Connect Serial' untuk melihat output.
3.  Klik 'Start Log' untuk mulai merekam data.
4.  Klik 'Pause Log' untuk menjeda (klik lagi untuk lanjut).
5.  Klik 'Stop & Save' untuk menyimpan rekaman data ke file Excel (.xlsx).
"""
        textbox.insert("1.0", help_text)
        textbox.configure(state="disabled")

    def on_port_select(self, port_description):
        if port_description == "-":
            self.selected_port = None
        else:
            self.selected_port = port_description.split(" ")[0]

    def refresh_ports(self):
        self.log_message("Mencari port COM...")
        ports = serial.tools.list_ports.comports()
        if not ports:
            self.com_port_menu.set("-")
            self.com_port_menu.configure(values=["-"])
            self.selected_port = None
            self.log_message("Port COM tidak ditemukan.")
        else:
            port_list = [f"{p.device} ({p.description})" for p in ports]
            self.com_port_menu.configure(values=port_list)
            self.com_port_menu.set(port_list[0])
            self.on_port_select(port_list[0])
            self.log_message(f"Ditemukan {len(port_list)} port.")

    def browse_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Binary files", "*.bin"), ("All files", "*.*")])
        if filepath:
            self.firmware_path.delete(0, "end")
            self.firmware_path.insert(0, filepath)
            self.log_message(f"File aplikasi dipilih: {filepath}")

    def log_message(self, msg):
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", msg + "\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")

    # --- [EDIT] Perbaikan Bug Tombol '...' ---
    def set_ui_state(self, state, is_flashing=True):
        if state == "disabled":
            if is_flashing:
                self.flash_button.configure(text="PROSES...", state="disabled")
                self.erase_button.configure(state="disabled")
                self.progress_bar.pack(fill="x", padx=10, pady=(0,10))
                self.progress_bar.start()
            self.com_port_menu.configure(state="disabled")
            self.baud_rate_menu.configure(state="disabled")
            self.refresh_ports_button.configure(state="disabled")
            self.browse_button.configure(state="disabled") # <-- BUG FIX
            if is_flashing:
                self.serial_connect_button.configure(state="disabled")
        else: 
            if is_flashing:
                self.flash_button.configure(text="UPLOAD FIRMWARE", state="normal")
                self.erase_button.configure(text="ERASE FLASH (HAPUS TOTAL)", state="normal")
                self.progress_bar.stop()
                self.progress_bar.pack_forget()
            self.com_port_menu.configure(state="normal")
            self.baud_rate_menu.configure(state="normal")
            self.refresh_ports_button.configure(state="normal")
            self.browse_button.configure(state="normal") # <-- BUG FIX
            if is_flashing:
                self.serial_connect_button.configure(state="normal")

    # --- [EDIT] 'run_command' kini mengembalikan (return_code, stderr_log) ---
    def run_command(self, command):
        self.log_message(f"Menjalankan: {' '.join(command)}")
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
        
        for line in process.stdout:
            self.log_message(line.strip())
            
        stderr_output = process.stderr.read()
        if stderr_output:
            self.log_message(f"LOG ERROR:\n{stderr_output.strip()}")
            
        process.wait()
        return process.returncode, stderr_output # <-- Perubahan di sini

    def start_flash_thread(self):
        if self.serial_running:
            self.log_message("Serial monitor aktif, otomatis disconnect...")
            self.disconnect_serial() 
            threading.Timer(0.5, lambda: threading.Thread(target=self.flash_firmware, daemon=True).start()).start()
        else:
            threading.Thread(target=self.flash_firmware, daemon=True).start()

    def start_erase_thread(self):
        if self.serial_running:
            self.log_message("Serial monitor aktif, otomatis disconnect...")
            self.disconnect_serial()
            threading.Timer(0.5, lambda: threading.Thread(target=self.erase_flash, daemon=True).start()).start()
        else:
            threading.Thread(target=self.erase_flash, daemon=True).start()

    # --- [EDIT] Menggunakan parser error ---
    def flash_firmware(self):
        if not self.selected_port:
            self.after(0, self.show_popup, "Error", "Silakan pilih COM Port.")
            return
        app_filepath = self.firmware_path.get()
        if not app_filepath:
            self.after(0, self.show_popup, "Error", "Silakan pilih file firmware APLIKASI (.bin)")
            return

        self.set_ui_state("disabled", is_flashing=True)
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")

        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))

            bootloader_path = os.path.join(base_path, "bootloader.bin")
            partitions_path = os.path.join(base_path, "partitions.bin")

            if not os.path.exists(bootloader_path):
                raise FileNotFoundError(f"bootloader.bin tidak ditemukan di {base_path}")
            if not os.path.exists(partitions_path):
                raise FileNotFoundError(f"partitions.bin tidak ditemukan di {base_path}")

            flash_baud = "921600" 
            command = [
                "python", "-m", "esptool", "--port", self.selected_port,
                "--baud", flash_baud, "--chip", "auto", "write_flash",
                "0x1000", bootloader_path, "0x8000", partitions_path,
                "0x10000", app_filepath
            ]

            return_code, stderr_log = self.run_command(command) # <-- Menangkap stderr
            
            if return_code == 0:
                self.log_message("\n--- UPLOAD BERHASIL! ---")
                self.after(0, self.show_popup, "Sukses", "Upload Firmware Berhasil!")
            else:
                self.log_message(f"\n--- UPLOAD GAGAL! ---")
                friendly_msg = self.parse_esptool_error(stderr_log) # <-- Menggunakan parser
                self.after(0, self.show_popup, "Upload Gagal", friendly_msg)

        except Exception as e:
            error_msg = f"Error Kritis: {e}"
            self.log_message(error_msg)
            friendly_msg = self.parse_esptool_error(str(e)) # <-- Menggunakan parser
            self.after(0, self.show_popup, "Error Kritis", friendly_msg)
        finally:
            self.set_ui_state("normal", is_flashing=True)

    # --- [EDIT] Menggunakan parser error ---
    def erase_flash(self):
        if not self.selected_port:
            self.after(0, self.show_popup, "Error", "Silakan pilih COM Port.")
            return

        self.set_ui_state("disabled", is_flashing=True)
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")

        try:
            flash_baud = "921600"
            command = [
                "python", "-m", "esptool", "--port", self.selected_port,
                "--baud", flash_baud, "erase_flash"
            ]
            
            return_code, stderr_log = self.run_command(command) # <-- Menangkap stderr

            if return_code == 0:
                self.log_message("\n--- ERASE BERHASIL! ---")
                self.after(0, self.show_popup, "Sukses", "Erase Flash Berhasil!")
            else:
                self.log_message(f"\n--- ERASE GAGAL! ---")
                friendly_msg = self.parse_esptool_error(stderr_log) # <-- Menggunakan parser
                self.after(0, self.show_popup, "Erase Gagal", friendly_msg)
                
        except Exception as e:
            error_msg = f"Error Kritis: {e}"
            self.log_message(error_msg)
            self.after(0, self.show_popup, "Error Kritis", error_msg)
        finally:
            self.set_ui_state("normal", is_flashing=True)

    # --- FUNGSI SERIAL MONITOR ---
    def log_serial(self, msg):
        self.serial_output.configure(state="normal")
        self.serial_output.insert("end", msg)
        self.serial_output.see("end")
        self.serial_output.configure(state="disabled")

    def clear_serial_output(self):
        self.serial_output.configure(state="normal")
        self.serial_output.delete("1.0", "end")
        self.serial_output.configure(state="disabled")

    def toggle_serial_connection(self):
        if self.ser and self.ser.is_open:
            self.disconnect_serial()
        else:
            self.connect_serial()

    def connect_serial(self):
        if not self.selected_port:
            self.after(0, self.show_popup, "Error Serial", "COM Port belum dipilih.")
            return
        
        try:
            baud = self.baud_rate_menu.get()
            self.ser = serial.Serial(self.selected_port, baud, timeout=1)
            self.serial_running = True
            
            self.read_thread = threading.Thread(target=self.serial_read_thread, daemon=True)
            self.read_thread.start()
            
            self.serial_connect_button.configure(text="Disconnect", fg_color="red", hover_color="darkred")
            self.log_serial(f"Terhubung ke {self.selected_port} @ {baud} bps.\n")
            
            # --- [EDIT] Perbaikan Bug Tombol '...' ---
            self.set_ui_state("disabled", is_flashing=False) 
            self.flash_button.configure(state="disabled")
            self.erase_button.configure(state="disabled")
            self.browse_button.configure(state="disabled") # <-- BUG FIX

        except Exception as e:
            self.log_serial(f"Gagal terhubung: {e}\n")
            self.after(0, self.show_popup, "Error Serial", f"Gagal terhubung: {e}")
            self.ser = None

    def disconnect_serial(self):
        self.serial_running = False
        if self.ser:
            self.ser.close()
            self.ser = None
        
        # Hentikan logging jika sedang berjalan
        if self.is_logging:
            self.stop_and_save_log()

        if self.serial_connect_button.winfo_exists():
            self.serial_connect_button.configure(text="Connect Serial", fg_color="#1F6AA5", hover_color="#144870")
            self.log_serial("\nKoneksi terputus.\n")
            
            # --- [EDIT] Perbaikan Bug Tombol '...' ---
            self.set_ui_state("normal", is_flashing=False)
            self.flash_button.configure(state="normal")
            self.erase_button.configure(state="normal")
            self.browse_button.configure(state="normal") # <-- BUG FIX

    # --- [EDIT] Ditambah logika logging ---
    def serial_read_thread(self):
        while self.serial_running:
            try:
                if self.ser and self.ser.in_waiting > 0:
                    data = self.ser.readline()
                    try:
                        decoded_data = data.decode('utf-8')
                        self.after(0, self.log_serial, decoded_data) 

                        # --- LOGIKA LOGGING BARU ---
                        if self.is_logging and not self.is_logging_paused:
                            # Tambahkan data ke buffer
                            self.log_data_buffer.append((datetime.now(), decoded_data))

                    except UnicodeDecodeError:
                        self.after(0, self.log_serial, f"[RAW: {data}]")
            except Exception as e:
                if self.serial_running:
                    self.after(0, self.log_serial, f"\nError baca serial: {e}\n")
                    self.after(0, self.disconnect_serial)
                break

    def send_serial_data(self, event=None):
        if self.ser and self.ser.is_open:
            data = self.serial_input.get()
            if data:
                try:
                    self.ser.write((data + '\n').encode('utf-8'))
                    self.serial_input.delete(0, "end") 
                except Exception as e:
                    self.after(0, self.show_popup, "Error Kirim", f"Gagal kirim data: {e}")
        else:
            self.after(0, self.show_popup, "Error Serial", "Tidak terhubung ke serial port.")

    # --- [FUNGSI BARU] Kontrol Logging ---
    def toggle_log_start(self):
        if openpyxl is None:
            self.after(0, self.show_popup, "Error Library", "Fitur Log gagal.\nLibrary 'openpyxl' tidak ditemukan.\n\nJalankan: pip install openpyxl")
            return
            
        if not self.serial_running:
            self.after(0, self.show_popup, "Error Log", "Silakan 'Connect Serial' terlebih dahulu.")
            return

        self.is_logging = True
        self.is_logging_paused = False
        self.log_data_buffer = [] # Kosongkan buffer
        
        self.log_start_button.configure(text="Logging...", state="disabled")
        self.log_pause_button.configure(text="Pause Log", state="normal")
        self.log_stop_button.configure(state="normal")
        self.log_serial("\n--- [LOGGING DIMULAI] ---\n")

    def toggle_log_pause(self):
        if not self.is_logging:
            return

        self.is_logging_paused = not self.is_logging_paused
        
        if self.is_logging_paused:
            self.log_pause_button.configure(text="Resume Log")
            self.log_serial("\n--- [LOGGING DIJEDA] ---\n")
        else:
            self.log_pause_button.configure(text="Pause Log")
            self.log_serial("\n--- [LOGGING DILANJUTKAN] ---\n")

    def stop_and_save_log(self):
        if not self.is_logging:
            return

        self.is_logging = False
        self.is_logging_paused = False
        self.log_serial("\n--- [LOGGING DIHENTIKAN] ---\n")

        # Reset tombol
        self.log_start_button.configure(text="Start Log", state="normal")
        self.log_pause_button.configure(text="Pause Log", state="disabled")
        self.log_stop_button.configure(text="Stop & Save", state="disabled")

        if not self.log_data_buffer:
            self.after(0, self.show_popup, "Info", "Tidak ada data log yang direkam.")
            return

        # Minta lokasi simpan file
        try:
            default_filename = f"esp32_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                title="Simpan Log Serial ke Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename
            )
            
            if not filepath: # Jika user menekan 'Cancel'
                self.log_message("Penyimpanan log dibatalkan.")
                return

            # Mulai proses simpan ke Excel (di thread agar tidak freeze)
            threading.Thread(target=self.save_to_excel, args=(filepath, self.log_data_buffer.copy()), daemon=True).start()
            self.log_data_buffer = [] # Kosongkan buffer

        except Exception as e:
            self.log_message(f"Error saat menyimpan Excel: {e}")
            self.after(0, self.show_popup, "Error Simpan", f"Gagal menyimpan file Excel:\n{e}")

    def save_to_excel(self, filepath, data_to_save):
        try:
            self.log_message(f"Menyimpan log ke: {filepath} ...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Serial Log"

            # Buat Header
            ws['A1'] = "Timestamp"
            ws['B1'] = "Data"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            
            # Sesuaikan lebar kolom
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 100

            # Tulis data
            for i, (timestamp, data) in enumerate(data_to_save, start=2):
                ws[f'A{i}'] = timestamp.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                ws[f'B{i}'] = data.strip() # Hapus newline/spasi ekstra

            wb.save(filepath)
            
            self.log_message("Log berhasil disimpan ke Excel.")
            self.after(0, self.show_popup, "Sukses", f"Log berhasil disimpan ke:\n{filepath}")

        except Exception as e:
            self.log_message(f"Error saat menyimpan Excel: {e}")
            self.after(0, self.show_popup, "Error Simpan", f"Gagal menyimpan file Excel:\n{e}")


if __name__ == "__main__":
    if openpyxl is None:
        # Tampilkan popup error jika library tidak ada
        root = ctk.CTk()
        root.withdraw() # Sembunyikan window utama
        app = App() # Panggil init untuk show_popup
        app.show_popup("Error Kritis: 'openpyxl' Hilang", 
                       "Library 'openpyxl' tidak ditemukan.\nFitur Log to Excel tidak akan jalan.\n\nSilakan tutup aplikasi ini, jalankan:\n\npip install openpyxl\n\nlalu jalankan lagi aplikasinya.")
        root.destroy()
    else:
        app = App()
        app.mainloop()